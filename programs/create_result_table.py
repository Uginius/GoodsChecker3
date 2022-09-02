import json
import math
import os
from utilites import get_last_dir
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

last_dir = get_last_dir()


def read_data_from_json():
    folder = 'htmls/' + last_dir
    json_files = [filename for filename in os.listdir(folder) if 'json' in filename]
    data = {}
    for filename in json_files:
        with open(f'{folder}/{filename}', 'r', encoding='utf8') as rf:
            data.update(json.load(rf))
    return data


def check_conditions(row):
    rating_cell, qt, need_cell = row[4], row[5].value, row[6]
    rating = rating_cell.value
    target = 4
    if rating is None:
        need = None
    else:
        if rating > target:
            need = 0 if qt > 1 else 1
        else:
            need = math.ceil(qt * abs(target - rating)) if rating else 2
    need_cell.value = need
    if need:
        need_cell.fill = PatternFill("solid", fgColor='d199ff')
    if rating_cell.value:
        if rating_cell.value == 0:
            rating_cell.fill = PatternFill("solid", fgColor='ffe4e1')
        elif rating_cell.value < 4:
            rating_cell.fill = PatternFill("solid", fgColor='E6B8B7')


def update_table(table, data):
    ws = table.active
    for order, row in enumerate(ws, 1):
        if order < 2:
            continue
        rating_cell = row[4]
        votes_cell = row[5]
        value = data.get(f'{order:03}')
        if not value:
            continue
        rating_cell.value, votes_cell.value = value
        check_conditions(row)
    table.save(f'results/{last_dir}.xlsx')


def update_result_table(xls_table):
    data = read_data_from_json()
    update_table(load_workbook(xls_table), data)
